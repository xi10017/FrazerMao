#!/usr/bin/env python3
"""Apply answer keys from TestParsing/Folder Answer Keys into famat_tests.json."""

from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("Install openpyxl: pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
FAMAT_PATH = ROOT / "src" / "data" / "famat_tests.json"
FOLDER_KEYS = Path(__file__).resolve().parent / "Folder Answer Keys"
STATS_STATES_CSV = Path(__file__).resolve().parent / "Statistics States Answers - new (1).csv"
IMPORT_PATH = Path(__file__).resolve().parent / "calculus_famat_import.json"

MONTH_MAP = {"jan": "January", "feb": "February", "mar": "March"}
REG_COL_RE = re.compile(r"^(Jan|Feb|Mar)(\d{2})(I|R|i)$", re.IGNORECASE)
REG_COL_ALT_RE = re.compile(r"^(Jan|Feb|Mar)(I|R)(\d{2})$", re.IGNORECASE)
YEAR_FLOAT_RE = re.compile(r"^(\d{4})(?:\.0)?$")
SECTION_HEADER_RE = re.compile(r"^[A-Za-z]")
SKIP_HEADERS = {"CALC", "PRECALC", "STATS", "Joey16", ""}
# Label columns that do not follow Jan15R naming but map to a specific test.
SPECIAL_REG_COLUMNS: dict[tuple[str, str], tuple[str, int, str, str]] = {
    ("Precalc Reg Season.xlsx", "PRECALC"): ("Alpha", 2016, "January", "Regional"),
}
# Spreadsheet gaps or known corrections (1-based question numbers).
MANUAL_ANSWER_OVERRIDES: dict[tuple, dict[int, str | list[str]]] = {
    ("Alg2", 2019, "March", "Invitational", "Individual"): {30: "B"},
}


@dataclass(frozen=True)
class AnswerKey:
    answers: list
    source: str
    file: str


ALL_ACCEPTED_ANSWERS = ["A", "B", "C", "D", "E"]


def parse_answer_cell(raw) -> str | list[str] | None:
    if raw is None:
        return None
    value = str(raw).strip()
    if not value:
        return None
    # e (0) -> E; strip parenthetical notes from answer cells.
    value = re.sub(r"\s*\([^)]*\)", "", value).strip()
    if not value:
        return None
    if value.lower() in {"throw", "x"}:
        return list(ALL_ACCEPTED_ANSWERS)
    # Thrown out / accept any response (e.g. "*", "* (ans)").
    if value.startswith("*"):
        return list(ALL_ACCEPTED_ANSWERS)
    if "/" in value:
        parts = [p.strip().upper() for p in value.split("/") if p.strip()]
        if len(parts) > 1:
            return parts
        value = parts[0] if parts else ""
    elif len(value) > 1 and value.replace(" ", "").isalpha():
        return [c.upper() for c in value if c.isalpha()]
    else:
        value = value.upper()
    if value in {"A", "B", "C", "D", "E"}:
        return value
    return None


def parse_reg_column(header: str) -> tuple[int, str, str] | None:
    if header in SKIP_HEADERS:
        return None
    match = REG_COL_RE.match(header)
    if match:
        month = MONTH_MAP[match.group(1).lower()]
        year = 2000 + int(match.group(2))
        test_type = "Invitational" if match.group(3).upper() == "I" else "Regional"
        return year, month, test_type

    match = REG_COL_ALT_RE.match(header)
    if match:
        month = MONTH_MAP[match.group(1).lower()]
        year = 2000 + int(match.group(3))
        test_type = "Invitational" if match.group(2).upper() == "I" else "Regional"
        return year, month, test_type
    return None


def normalize_source(header: str) -> str:
    match = REG_COL_RE.match(header)
    if match and match.group(3) == "i":
        return header[:-1] + "i"
    return header


def load_sheet_rows(path: Path, sheet_name: str | None = None) -> list[list]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name or wb.sheetnames[0]]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def trim_answers(answers: list) -> list:
    trimmed = list(answers)
    while trimmed and trimmed[-1] is None:
        trimmed.pop()
    return trimmed


def apply_manual_overrides(answers: list, key: tuple) -> list:
    overrides = MANUAL_ANSWER_OVERRIDES.get(key)
    if not overrides:
        return answers
    result = list(answers)
    for question, answer in overrides.items():
        while len(result) < question:
            result.append(None)
        result[question - 1] = answer
    return trim_answers(result)


def apply_manual_overrides_to_lookup(lookup: dict[tuple, AnswerKey]) -> None:
    for key, question_overrides in MANUAL_ANSWER_OVERRIDES.items():
        if key in lookup:
            existing = lookup[key]
            lookup[key] = AnswerKey(
                answers=apply_manual_overrides(existing.answers, key),
                source=existing.source,
                file=existing.file,
            )
        else:
            answers: list = []
            lookup[key] = AnswerKey(
                answers=apply_manual_overrides(answers, key),
                source="manual_override",
                file="manual_override",
            )


def parse_reg_season_file(
    path: Path,
    division: str,
    sheet_name: str | None = None,
) -> dict[tuple, AnswerKey]:
    rows = load_sheet_rows(path, sheet_name)
    if not rows:
        return {}

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    lookup: dict[tuple, AnswerKey] = {}

    for col_idx, header in enumerate(headers):
        meta = parse_reg_column(header)
        special = SPECIAL_REG_COLUMNS.get((path.name, header))
        if special:
            division_override, year, month, test_type = special
            answers: list = []
            for row in rows[1:]:
                cell = row[col_idx] if col_idx < len(row) else None
                answers.append(parse_answer_cell(cell))
            if any(a is not None for a in answers):
                key = (division_override, year, month, test_type, "Individual")
                lookup[key] = AnswerKey(
                    answers=trim_answers(answers),
                    source=header,
                    file=str(path.relative_to(FOLDER_KEYS.parent)),
                )
            continue
        if not meta:
            continue
        year, month, test_type = meta
        answers: list = []
        for row in rows[1:]:
            cell = row[col_idx] if col_idx < len(row) else None
            answers.append(parse_answer_cell(cell))
        if not any(a is not None for a in answers):
            continue
        key = (division, year, month, test_type, "Individual")
        lookup[key] = AnswerKey(
            answers=trim_answers(answers),
            source=normalize_source(header),
            file=str(path.relative_to(FOLDER_KEYS.parent)),
        )
    return lookup


def parse_format_from_header(header: str) -> str | None:
    h = header.lower()
    if h.startswith("unplug"):
        return "Unplugged Individual"
    if h.startswith(("ind", "indiv")):
        return "Individual"
    return None


def is_topic_section_header(header: str) -> bool:
    """Non-Individual test section label (Apps13, Quad08, Prob16, etc.)."""
    if not header or YEAR_FLOAT_RE.match(header):
        return False
    if parse_format_from_header(header):
        return False
    return bool(SECTION_HEADER_RE.match(header))


def parse_base_year_from_header(header: str) -> int | None:
    digits = re.findall(r"\d+", header)
    if not digits:
        return None
    year = int(digits[0])
    return year if year >= 1900 else 2000 + year


def parse_states_nationals_file(
    path: Path,
    division: str,
    test_type: str,
    sheet_name: str | None = None,
) -> dict[tuple, AnswerKey]:
    rows = load_sheet_rows(path, sheet_name)
    if not rows:
        return {}

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    lookup: dict[tuple, AnswerKey] = {}
    current_format: str | None = None
    current_base_year: int | None = None

    for col_idx, header in enumerate(headers):
        if is_topic_section_header(header):
            current_format = None
            current_base_year = None
            continue

        fmt = parse_format_from_header(header)
        if fmt:
            current_format = fmt
            current_base_year = parse_base_year_from_header(header)
            year = current_base_year
        elif YEAR_FLOAT_RE.match(header) and current_format and current_base_year:
            year = int(float(header))
        else:
            continue

        if not current_format or year is None:
            continue

        answers: list = []
        for row in rows[1:]:
            cell = row[col_idx] if col_idx < len(row) else None
            answers.append(parse_answer_cell(cell))
        if not any(a is not None for a in answers):
            continue

        key = (division, year, None, test_type, current_format)
        lookup[key] = AnswerKey(
            answers=trim_answers(answers),
            source=f"{header}:{test_type}{year}",
            file=str(path.relative_to(FOLDER_KEYS.parent)),
        )
    return lookup


def load_stats_states_csv() -> dict[tuple, AnswerKey]:
    if not STATS_STATES_CSV.exists():
        return {}

    lookup: dict[tuple, AnswerKey] = {}
    with STATS_STATES_CSV.open(newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    headers = rows[0]
    current_format: str | None = None
    current_base_year: int | None = None

    for col_idx, header in enumerate(headers):
        if is_topic_section_header(header):
            current_format = None
            current_base_year = None
            continue

        fmt = parse_format_from_header(header)
        if fmt:
            current_format = fmt
            current_base_year = parse_base_year_from_header(header)
            year = current_base_year
        elif header.isdigit() and current_format and current_base_year:
            year = int(header)
        else:
            continue

        answers: list = []
        for row in rows[1:]:
            cell = row[col_idx] if col_idx < len(row) else ""
            answers.append(parse_answer_cell(cell))
        if not any(a is not None for a in answers):
            continue

        key = ("Stats", year, None, "States", current_format)
        lookup[key] = AnswerKey(
            answers=trim_answers(answers),
            source=header if fmt else f"States{year}",
            file=STATS_STATES_CSV.name,
        )
    return lookup


def load_all_answer_keys() -> dict[tuple, AnswerKey]:
    lookup: dict[tuple, AnswerKey] = {}

    def merge(source: dict[tuple, AnswerKey]) -> None:
        for key, value in source.items():
            lookup[key] = value

    reg_files = [
        (FOLDER_KEYS / "regular season" / "Precalc Reg Season.xlsx", "Alpha", "Sheet2"),
        (FOLDER_KEYS / "regular season" / "Stats Reg Season.xlsx", "Stats", "Sheet2"),
        (FOLDER_KEYS / "regular season" / "Calc Reg Season.xlsx", "Mu", "Sheet1"),
        (FOLDER_KEYS / "regular season" / "Alg 1 Reg Season.xlsx", "Alg1", "Sheet1"),
        (FOLDER_KEYS / "regular season" / "Alg 2 Reg Season.xlsx", "Alg2", "Sheet1"),
        (FOLDER_KEYS / "regular season" / "Geo Reg Season.xlsx", "Geo", "Sheet2"),
    ]
    for path, division, sheet in reg_files:
        if path.exists():
            merge(parse_reg_season_file(path, division, sheet))

    states_files = [
        (FOLDER_KEYS / "states" / "Alpha States Answers.xlsx", "Alpha"),
        (FOLDER_KEYS / "states" / "Mu States Answers.xlsx", "Mu", "Sheet2"),
        (FOLDER_KEYS / "states" / "Theta States Answers.xlsx", "Theta", "Old ver."),
        (FOLDER_KEYS / "states" / "Theta States Answers.xlsx", "Theta", "new"),
    ]
    for item in states_files:
        path = item[0]
        division = item[1]
        sheet = item[2] if len(item) > 2 else None
        if path.exists():
            merge(parse_states_nationals_file(path, division, "States", sheet))

    nats_files = [
        (FOLDER_KEYS / "nationals" / "Alpha Nats Answers.xlsx", "Alpha"),
        (FOLDER_KEYS / "nationals" / "Mu Answers.xlsx", "Mu", "2013-Present"),
        (FOLDER_KEYS / "nationals" / "Theta Nats Answers.xlsx", "Theta", "Sheet1"),
    ]
    for item in nats_files:
        path = item[0]
        division = item[1]
        sheet = item[2] if len(item) > 2 else None
        if path.exists():
            merge(parse_states_nationals_file(path, division, "Nationals", sheet))

    merge(load_stats_states_csv())
    apply_manual_overrides_to_lookup(lookup)
    return lookup


def solution_key(entry: dict) -> tuple:
    return (
        entry.get("division"),
        entry.get("year"),
        entry.get("month"),
        entry.get("test_type"),
        entry.get("format"),
    )


def get_test_id(entry: dict) -> str:
    month_part = f"-{entry['month']}" if entry.get("month") else ""
    return (
        f"{entry['year']}{month_part}-{entry['division']}-{entry['test_type']}-{entry['format']}"
        .lower()
        .replace(" ", "-")
        .replace("(", "")
        .replace(")", "")
    )


def score_entry(entry: dict) -> int:
    name = entry.get("name", "").lower()
    score = 0
    if re.search(r"\d+\s+t_", entry.get("name", ""), re.I):
        score += 100
    if re.search(r"\d+\s+s_", entry.get("name", ""), re.I):
        score += 100
    if "calculus individual" in name or "calc indiv" in name:
        score += 50
    if "mu open" in name:
        score -= 80
    if any(x in name for x in ("errata", " err", "coverpage", "err.doc")):
        score -= 100
    if re.search(r"\bE_", entry.get("name", ""), re.I):
        score -= 200
    if re.search(r"\bT_", entry.get("name", ""), re.I):
        score += 80
    return score


def pick_best(entries: list[dict]) -> dict:
    return max(entries, key=score_entry)


def add_mu_tests_with_keys(data: list[dict], lookup: dict[tuple, AnswerKey]) -> list[dict]:
    if not IMPORT_PATH.exists():
        return data

    imported = json.loads(IMPORT_PATH.read_text(encoding="utf-8"))
    existing_test_ids = {get_test_id(e) for e in data if e.get("document_type") == "Test"}
    existing_keys = {solution_key(e) for e in data}

    mu_candidates: dict[tuple, dict[str, list[dict]]] = {}
    for entry in imported:
        if entry.get("division") != "Mu" or entry.get("format") != "Individual":
            continue
        if entry.get("test_type") not in {"Regional", "Invitational"}:
            continue
        key = solution_key(entry)
        if key not in lookup:
            continue
        mu_candidates.setdefault(key, {"Test": [], "Solution": []})
        mu_candidates[key][entry["document_type"]].append(entry)

    additions: list[dict] = []
    for key, groups in sorted(mu_candidates.items()):
        if key in existing_keys:
            continue
        if not groups["Test"] or not groups["Solution"]:
            continue
        test = pick_best(groups["Test"])
        solution = dict(pick_best(groups["Solution"]))
        answer = lookup[key]
        solution["answers"] = answer.answers
        solution["csv_validated"] = True
        solution["csv_source"] = answer.source
        solution["answer_key_file"] = answer.file
        if get_test_id(test) in existing_test_ids:
            continue
        additions.extend([test, solution])

    return data + additions


def main() -> None:
    lookup = load_all_answer_keys()
    data = json.loads(FAMAT_PATH.read_text(encoding="utf-8"))
    data = add_mu_tests_with_keys(data, lookup)

    updated = 0
    missing: list[tuple] = []
    changed: list[str] = []

    for entry in data:
        if entry.get("document_type") != "Solution":
            continue
        key = solution_key(entry)
        answer = lookup.get(key)
        if not answer:
            if entry.get("answers"):
                continue
            missing.append(key)
            continue

        old_answers = entry.get("answers")
        entry["answers"] = answer.answers
        entry["csv_validated"] = True
        entry["csv_source"] = answer.source
        entry["answer_key_file"] = answer.file
        updated += 1
        if old_answers and old_answers != answer.answers:
            changed.append(
                f"{key} ({entry.get('csv_source')} -> {answer.source})"
            )

    data.sort(
        key=lambda e: (
            e.get("division") or "",
            -(e.get("year") or 0),
            e.get("month") or "",
            e.get("test_type") or "",
            e.get("document_type") or "",
            e.get("name") or "",
        )
    )

    tests = [e for e in data if e.get("document_type") == "Test"]
    ids = [get_test_id(t) for t in tests]
    dupes = sorted({i for i in ids if ids.count(i) > 1})
    if dupes:
        raise SystemExit(f"Duplicate test IDs: {dupes}")

    solutions = [e for e in data if e.get("document_type") == "Solution"]
    without = [solution_key(s) for s in solutions if not s.get("answers")]

    FAMAT_PATH.write_text(json.dumps(data, indent=4) + "\n", encoding="utf-8")

    print(f"Loaded {len(lookup)} answer key columns from folder + CSV")
    print(f"Updated {updated} solution entries")
    print(f"Solutions without answer keys: {len(without)}")
    print(f"Total entries: {len(data)} ({len(tests)} tests)")
    if changed:
        print(f"\nAnswer changes detected ({len(changed)}):")
        for line in changed[:20]:
            print(f"  {line}")
        if len(changed) > 20:
            print(f"  ... and {len(changed) - 20} more")
    if without:
        print("\nMissing answer keys for:")
        for key in without[:25]:
            print(f"  {key}")
        if len(without) > 25:
            print(f"  ... and {len(without) - 25} more")

    unused = sorted(
        set(lookup) - {solution_key(s) for s in solutions if s.get("answers")},
        key=lambda k: (k[0] or "", -(k[1] or 0), k[2] or "", k[3] or "", k[4] or ""),
    )
    app_keys = {solution_key(s) for s in solutions}
    unused_in_app = [k for k in unused if k[0] in {"Alpha", "Stats", "Mu", "Alg1", "Alg2", "Geo", "Theta"}]
    if unused_in_app:
        print(f"\nAnswer keys with no matching test in catalog ({len(unused_in_app)}):")
        for key in unused_in_app[:15]:
            print(f"  {key}")
        if len(unused_in_app) > 15:
            print(f"  ... and {len(unused_in_app) - 15} more (mostly tests without PDFs)")


if __name__ == "__main__":
    main()
